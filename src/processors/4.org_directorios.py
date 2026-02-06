import os
import shutil
import pandas as pd
import re
import sys

# Agregar raíz del proyecto al path
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.path.insert(0, PROJECT_ROOT)

from src.utils.buscar_proyecto_maestra import buscar_proyecto, cargar_excel
import glob
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from pathlib import Path
import io

# ==================== CONFIGURACIÓN UTF-8 PARA WINDOWS ====================
# Forzar UTF-8 en stdout y stderr para evitar errores de codificación
if sys.platform == 'win32':
    if sys.stdout.encoding != 'utf-8':
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    if sys.stderr.encoding != 'utf-8':
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
# ==========================================================================


# Verificar si se pasó un parámetro desde la línea de comandos
if len(sys.argv) > 1:
    # Usar el primer argumento como ruta base
    BASE_PATH = sys.argv[1]
    print(f"Usando directorio proporcionado: {BASE_PATH}")
else:
    # Usar una ruta predetermincada si no se proporciona ningún argumento
    BASE_PATH = "inboxFacturas/RV_ Facturación Perú 16012025"
    print(f"Ningún directorio proporcionado. Usando ruta predeterminada: {BASE_PATH}")

# Archivos de entrada ajustados dinámicamente
EXCEL_FILE = os.path.join(BASE_PATH, "3.file_table_xml.xlsx")
CSV_FILE = os.path.join(BASE_PATH, "tabla_1.csv")
EXISTING_EXCEL = r"D:\facturas_bot\Registros Historico\3.Historicov2.xlsx"  # Ruta al Excel existente donde también se agregarán los datos
PATH_DIRECTORY_FILES= r"D:\facturas_bot\Documentos de Facturación SSFF"
DIRECTORY_FILE_XML = f"{BASE_PATH}/comprobantes_XML"
NEW_EXCEL = Path(DIRECTORY_FILE_XML).parent / "3.file_table_xml.xlsx"  # Ruta para el Excel nuevo

print(f"Ruta del archivo Excel: {EXCEL_FILE}")
print(f"Ruta del archivo CSV: {CSV_FILE}")

# Verificar si los archivos existen
if not os.path.exists(EXCEL_FILE):
    print(f" ERROR: El archivo Excel no existe en {EXCEL_FILE}")
    sys.exit(1)

if not os.path.exists(CSV_FILE):
    print(f" ERROR: El archivo CSV no existe en {CSV_FILE}")
    sys.exit(1)

# Buscar carpetas de comprobantes en la ruta base
CARPETAS_COMPROBANTES = [
    d for d in os.listdir(BASE_PATH) if os.path.isdir(os.path.join(BASE_PATH, d)) and (
        d.startswith("comprobantes_XML") or d.startswith("comprobantes_PDF") or d.startswith("comprobantes_CDR")
    )
]

print(f"Carpetas de comprobantes encontradas: {CARPETAS_COMPROBANTES}")

# Carpeta base donde se crearán las estructuras
BASE_DIR = os.path.join(BASE_PATH, "Organizado")
os.makedirs(BASE_DIR, exist_ok=True)

# Archivo de log
LOG_FILE = os.path.join(BASE_DIR, "log.xlsx")

# Cargar los datos
df_excel = pd.read_excel(EXCEL_FILE, dtype=str).fillna("")
# Agregar la nueva columna "EN MAESTRA" al dataframe
df_excel["EN MAESTRA"] = ""

df_csv = pd.read_csv(CSV_FILE, delimiter='|', header=None, dtype=str,on_bad_lines='warn').fillna("")

# Lista para guardar logs
logs = []

def identificar_columnas_csv(df):
    """
    Identifica dinámicamente las columnas del CSV basadas en patrones:
    - Columna del patrón XX-XXXX-XXXXX (para buscar comprobantes)
    - Columna con exactamente 6 caracteres alfanuméricos (para el proyecto/carpeta)
    
    Retorna: (índice_columna_comprobante, índice_columna_proyecto)
    """
    indice_columna_comprobante = None
    indice_columna_proyecto = None
    
    # Expresión regular para el patrón XX-XXXX-XXXXX (manteniendo los dos guiones como solicitaste)
    patron_comprobante = re.compile(r'^\d{2}-[A-Z0-9]{1,4}--\d{7}$')
    
    # Expresión regular para exactamente 6 caracteres alfanuméricos
    patron_proyecto = re.compile(r'^[A-Za-z0-9]{6}$')
    
    print(f"Patrones a buscar:")
    print(f"- Comprobante: {patron_comprobante.pattern}")
    print(f"- Proyecto: {patron_proyecto.pattern} (6 caracteres alfanuméricos)")
    print(f"\nAnalizando DataFrame con {len(df)} filas y {len(df.columns)} columnas")
    
    # Examinar las primeras filas para determinar los patrones
    num_filas_analizar = min(10, len(df))
    print(f"Se analizarán las primeras {num_filas_analizar} filas")
    
    # Para cada columna, verificar si cumple con alguno de los patrones
    for col_idx in range(len(df.columns)):
        nombre_columna = df.columns[col_idx] if hasattr(df, 'columns') and len(df.columns) > col_idx else f'Columna {col_idx}'
        print(f"\n--- Analizando columna {col_idx}: {nombre_columna} ---")
        valores_comprobante = 0
        valores_proyecto = 0
        
        for idx in range(num_filas_analizar):
            if idx < len(df):
                valor = str(df.iloc[idx, col_idx]).strip()
                
                # Verificar si es NaN o valor vacío antes de evaluar patrones
                if valor and valor.lower() != 'nan':
                    if patron_comprobante.match(valor):
                        valores_comprobante += 1
                        print(f"    OK  Coincide con patrón de comprobante")
                    else:
                        print(f"    ERROR No coincide con patrón de comprobante")
                    
                    if patron_proyecto.match(valor):
                        valores_proyecto += 1
                        print(f"    OK  Coincide con patrón de proyecto (6 caracteres alfanuméricos)")
                    else:
                        print(f"    ERROR No coincide con patrón de proyecto")
                        # Analizar por qué no coincide con el patrón de proyecto
                        if len(valor) != 6:
                            print(f"      - > La longitud es {len(valor)}, se esperaban 6 caracteres")
                        if not re.match(r'^[A-Za-z0-9]*$', valor):
                            print(f"      - > Contiene caracteres que no son alfanuméricos")
                else:
                    print(f"    ! Valor vacío o NaN - ignorado")
        
        print(f"  Resumen columna {col_idx}:")
        print(f"    - Valores que coinciden con patrón de comprobante: {valores_comprobante}/{num_filas_analizar}")
        print(f"    - Valores que coinciden con patrón de proyecto: {valores_proyecto}/{num_filas_analizar}")
        
        # Si más del 50% de los valores cumplen el patrón, asumimos que es la columna correcta
        if valores_comprobante > num_filas_analizar / 2 and indice_columna_comprobante is None:
            indice_columna_comprobante = col_idx
            print(f"   Columna de comprobantes identificada: {col_idx} (patrón XX-XXXX-XXXXX)")
        
        if valores_proyecto > num_filas_analizar / 2 and indice_columna_proyecto is None:
            indice_columna_proyecto = col_idx
            print(f"   Columna de proyecto identificada: {col_idx} (patrón 6 caracteres alfanuméricos)")
    
    print("\n=== RESULTADOS FINALES ===")
    # Si no se encontró alguna columna, usar valores predeterminados y avisar
    if indice_columna_comprobante is None:
        indice_columna_comprobante = 0
        print(f" No se pudo identificar automáticamente la columna de comprobantes. Usando predeterminado: 0")
    else:
        print(f" Columna de comprobantes: {indice_columna_comprobante}")
    
    if indice_columna_proyecto is None:
        indice_columna_proyecto = 3
        print(f" No se pudo identificar automáticamente la columna de proyecto. Usando predeterminado: 3")
    else:
        print(f" Columna de proyecto: {indice_columna_proyecto}")
    
    return indice_columna_comprobante, indice_columna_proyecto

# Identificar las columnas del CSV
columna_comprobante, columna_proyecto = identificar_columnas_csv(df_csv)
print(f"Se usará la columna {columna_comprobante} para buscar comprobantes y la columna {columna_proyecto} para obtener el proyecto")

def create_excel_with_headers(output_file):
    """
    Crea un archivo Excel con los encabezados especificados y formato
    """
    headers = [
        "Cliente", "RUC", "Proyecto", "Empresa INDRA/MPS/TCN", "RUC2", 
        "N° de Comprobante", "Fecha de Envío", "Divisa", "Tipo de Impuesto",
        "Condición de pago", "Valor Venta", "IGV (18%)", "TOTAL", 
        "OC-OS", "Número de Recepción (NR-CR)", "Descripción (Primera Fila)",
        "ENVAR CORREO", "ESTADO", "EN MAESTRA"  # Añadida la nueva columna
    ]
    
    # Crear DataFrame vacío con las columnas
    df = pd.DataFrame(columns=headers)
    
    # Guardar a Excel
    df.to_excel(output_file, index=False, sheet_name='Facturas')
    
    # Dar formato con openpyxl
    wb = openpyxl.load_workbook(output_file)
    ws = wb['Facturas']
    
    # Estilo para encabezados
    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    value_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    description_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # Aplicar bordes a todas las celdas
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Formato para encabezados
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        
        # Aplicar color según tipo de columna
        if header in ["Valor Venta", "IGV (18%)", "TOTAL","ENVAR CORREO","ESTADO", "EN MAESTRA"]:
            cell.fill = value_fill
        elif header in ["Descripción (Primera Fila)"]:
            cell.fill = description_fill
        else:
            cell.fill = header_fill
    
    # Ajustar ancho de columnas
    for col_num, _ in enumerate(headers, 1):
        col_letter = openpyxl.utils.get_column_letter(col_num)
        if col_num in [1, 3, 4, 16]:  # Nombres, proyectos y descripción más anchos
            ws.column_dimensions[col_letter].width = 25
        elif col_num in [6, 7, 9, 10]:  # Fechas, condiciones con ancho medio
            ws.column_dimensions[col_letter].width = 15
        else:
            ws.column_dimensions[col_letter].width = 12
    
    wb.save(output_file)
    return output_file

def append_to_existing_excel(df_actualizado):
    """
    Agrega o actualiza datos del DataFrame actualizado al Excel histórico
    Solo incluye registros con 'OK' en la columna EN MAESTRA
    
    Args:
        df_actualizado: DataFrame con los datos actualizados incluyendo la columna EN MAESTRA
    """
    existing_excel = EXISTING_EXCEL
    print(f"\n Iniciando actualización del Excel histórico: {existing_excel}")

    try:
        # Filtrar solo los registros con "OK" en la columna EN MAESTRA
        df_filtrado = df_actualizado[df_actualizado['EN MAESTRA'] == 'OK'].copy()
        print(f" Filtrando registros con 'OK' en EN MAESTRA: {len(df_filtrado)} de {len(df_actualizado)} registros")
        
        # Si no hay registros con OK, informar y salir
        if len(df_filtrado) == 0:
            print(" No hay registros con 'OK' en la columna EN MAESTRA. No se actualizará el Excel histórico.")
            return None
        
        # El resto de la función sigue igual, pero trabajando con df_filtrado en lugar de df_actualizado
        
        # Verificar si el archivo Excel existe
        if not os.path.exists(existing_excel):
            print(f" El archivo Excel histórico {existing_excel} no existe. Creando una copia...")

            # Crear directorio si no existe
            os.makedirs(os.path.dirname(existing_excel), exist_ok=True)
            
            # Crear un nuevo Excel con encabezados
            output_file = create_excel_with_headers(existing_excel)
            print(f" Nuevo archivo histórico creado en: {existing_excel}")
            
            # Ahora agregar los datos al Excel recién creado
            df_filtrado.to_excel(existing_excel, index=False, sheet_name='Facturas')
            print(f" Datos agregados al nuevo Excel histórico")
            return existing_excel

        # Si el archivo existe, cargar datos
        print(f" Leyendo Excel histórico existente...")
        df_historico = pd.read_excel(existing_excel)
        print(f" Excel histórico cargado. Contiene {len(df_historico)} registros.")

        # Contador para seguimiento
        actualizados = 0
        agregados = 0
        
        # Para cada factura en el dataframe filtrado
        print(f" Procesando {len(df_filtrado)} registros del dataframe filtrado...")
        
        # Crear un nuevo dataframe para almacenar los registros a agregar
        nuevos_registros = []
        
        for _, row in df_filtrado.iterrows():
            # Verificar si la factura ya existe (por número de comprobante)
            comprobante = row.get('N° de Comprobante', '')
            
            # Buscar el comprobante en el histórico
            if comprobante and 'N° de Comprobante' in df_historico.columns:
                # Verificar si existe
                mask = df_historico['N° de Comprobante'] == comprobante
                if mask.any():
                    # Si existe, actualizar la fila correspondiente con la nueva columna EN MAESTRA
                    indice = df_historico.loc[mask].index[0]
                    df_historico.loc[indice, 'EN MAESTRA'] = row.get('EN MAESTRA', '')
                    actualizados += 1
                else:
                    # Si no existe, agregar la fila completa al dataframe histórico
                    nuevos_registros.append(row)
                    agregados += 1
            else:
                # Si el dataframe histórico no tiene la columna o está vacío
                nuevos_registros.append(row)
                agregados += 1
        
        # Agregar los nuevos registros al dataframe histórico
        if nuevos_registros:
            df_nuevos = pd.DataFrame(nuevos_registros)
            df_historico = pd.concat([df_historico, df_nuevos], ignore_index=True)
        
        # Asegurarse de que el dataframe histórico tenga la columna EN MAESTRA
        if 'EN MAESTRA' not in df_historico.columns:
            df_historico['EN MAESTRA'] = ""
        
        # Guardar el Excel actualizado
        df_historico.to_excel(existing_excel, index=False)
        
        print(f"\n Resumen de actualización del Excel histórico:")
        print(f"    Registros actualizados: {actualizados}")
        print(f"    Registros agregados: {agregados}")
        print(f"    Total de registros en el histórico: {len(df_historico)}")
        print(f"    Excel histórico guardado en: {existing_excel}")
        
        return existing_excel

    except Exception as e:
        print(f" Error al agregar datos al Excel histórico: {str(e)}")
        import traceback
        print(traceback.format_exc())  # Imprime el stack trace completo
        return None
    
def transformar_codigo(comprobante):
    """Convierte el formato de 'F001-036500' a '01-F001--0036500' para buscar en el CSV."""
    partes = comprobante.split("-")
    if len(partes) == 2:
        serie, numero = partes
        numero = numero.zfill(7)
        return f"01-{serie}--{numero}"
    return comprobante

def transformar_codigo_3(comprobante):
    """Convierte el formato de 'F001-036500' a '01-F001--0036500' para buscar en el CSV."""
    partes = comprobante.split("-")
    if len(partes) == 2:
        serie, numero = partes
        numero = numero.zfill(7)
        return f"03-{serie}--{numero}"
    return comprobante

def buscar_y_copiar_archivos_OC(texto_buscar, carpeta_origen, carpeta_destino, incluir_subcarpetas=True):
    archivos_encontrados = []
    
    print(f" Iniciando búsqueda en carpeta origen: {carpeta_origen}")
    print(f" Carpeta destino: {carpeta_destino}")
    print(f" Texto a buscar: '{texto_buscar}'")
    print(f" Búsqueda en subcarpetas: {'Sí' if incluir_subcarpetas else 'No'}")
    
    # Crear carpeta destino si no existe
    if not os.path.exists(carpeta_destino):
        print(f" Creando carpeta destino: {carpeta_destino}")
        os.makedirs(carpeta_destino, exist_ok=True)
    else:
        print(f" Carpeta destino ya existe: {carpeta_destino}")
    
    print(f" Buscando archivos con texto '{texto_buscar}' en {carpeta_origen}")
    
    # Definir patrón de búsqueda
    if incluir_subcarpetas:
        patron_busqueda = os.path.join(carpeta_origen, "**", f"*{texto_buscar}*")
        print(f" Patrón de búsqueda (con subcarpetas): {patron_busqueda}")
        archivos = glob.glob(patron_busqueda, recursive=True)
    else:
        patron_busqueda = os.path.join(carpeta_origen, f"*{texto_buscar}*")
        print(f" Patrón de búsqueda (sin subcarpetas): {patron_busqueda}")
        archivos = glob.glob(patron_busqueda)
    
    print(f" Total de elementos encontrados (antes de filtrar): {len(archivos)}")
    
    # Filtrar solo archivos (no carpetas)
    archivos_antes = len(archivos)
    archivos = [a for a in archivos if os.path.isfile(a)]
    print(f" Elementos filtrados (solo archivos): {len(archivos)} de {archivos_antes}")
    
    if not archivos:
        print(f" No se encontraron archivos con '{texto_buscar}' en el nombre")
        return []
    
    print(f" Se encontraron {len(archivos)} archivos")
    
    # Listar archivos encontrados
    print(" Lista de archivos encontrados:")
    for i, archivo in enumerate(archivos, 1):
        print(f"   {i}. {archivo}")
    
    # Copiar cada archivo encontrado
    print(" Iniciando proceso de copia...")
    copiados = 0
    errores = 0
    
    for archivo_origen in archivos:
        nombre_archivo = os.path.basename(archivo_origen)
        archivo_destino = os.path.join(carpeta_destino, nombre_archivo)
        
        print(f"    Copiando: {nombre_archivo}")
        print(f"      De: {archivo_origen}")
        print(f"      A:  {archivo_destino}")
        
        try:
            shutil.copy2(archivo_origen, archivo_destino)
            print(f"    Copiado exitosamente: {nombre_archivo}")
            archivos_encontrados.append(archivo_origen)
            copiados += 1
        except Exception as e:
            print(f"    Error al copiar {archivo_origen}: {str(e)}")
            errores += 1
    
    print("\n Resumen:")
    print(f"    Carpeta origen: {carpeta_origen}")
    print(f"    Carpeta destino: {carpeta_destino}")
    print(f"    Texto buscado: '{texto_buscar}'")
    print(f"    Archivos encontrados: {len(archivos)}")
    print(f"    Archivos copiados: {copiados}")
    print(f"    Errores: {errores}")
    
    return archivos_encontrados

def generar_variantes_nombre(archivo_nombre):
    """
    Genera variantes del nombre de archivo:
    - Original: 20123456789-01-F001-036500
    - Sin ceros: 20123456789-01-F001-36500
    """
    variantes = [archivo_nombre]  # Siempre incluir el original
    
    # Buscar el patrón después del último guion y remover ceros a la izquierda
    partes = archivo_nombre.rsplit('-', 1)  # Separar por el último guion
    if len(partes) == 2:
        prefijo, numero = partes
        # Remover ceros a la izquierda del número
        numero_sin_ceros = numero.lstrip('0') or '0'  # Mantener al menos un 0 si todo son ceros
        if numero != numero_sin_ceros:
            variante_sin_ceros = f"{prefijo}-{numero_sin_ceros}"
            variantes.append(variante_sin_ceros)
            print(f"    Generada variante sin ceros: {variante_sin_ceros}")
    
    return variantes
    
def buscar_y_copiar_archivos(archivo_nombre, destino):
    """Busca en las carpetas de comprobantes y copia los archivos encontrados."""
    archivos_encontrados = []
    
    print(f" Buscando archivos para: {archivo_nombre}")
    
    # Generar variantes del nombre (con y sin ceros)
    variantes = generar_variantes_nombre(archivo_nombre)
    print(f"    Variantes a buscar: {variantes}")

    for carpeta in CARPETAS_COMPROBANTES:
        print("carpeta: " + carpeta)
        carpeta_path = os.path.join(BASE_PATH, carpeta)
        
        if not os.path.exists(carpeta_path): 
            print(f" Carpeta no encontrada: {carpeta_path}")
            continue
        
        for ext in [".xml", ".pdf", ".cdr"]:
            if ext.replace(".", "").upper() in carpeta:
                # Probar con cada variante del nombre
                for variante in variantes:
                    archivo_actual = variante
                    
                    print("ext:" + ext)
                    if ext == ".cdr":
                        ext_buscar = ".xml"
                        archivo_actual = "R-" + archivo_actual
                    else:
                        ext_buscar = ext
                    
                    archivo_origen = os.path.join(carpeta_path, archivo_actual + ext_buscar)
                    print(f"    Verificando: {archivo_origen}")

                    if os.path.exists(archivo_origen):
                        print(f"    Archivo encontrado y copiado: {archivo_origen} - > {destino}")
                        shutil.copy(archivo_origen, destino)
                        archivos_encontrados.append(archivo_origen)
                        break  # Si encontró con esta variante, no probar las demás
                    else:
                        print(f"    Archivo no encontrado: {archivo_origen}")

    return archivos_encontrados

# Procesar cada fila del Excel
for idx, row in df_excel.iterrows():
    proyecto = row["Proyecto"].replace(".","")
    
    # Verificar si el proyecto está vacío
    if not proyecto or proyecto.strip() == "":
        df_excel.at[idx, "EN MAESTRA"] = "NO"
        print("Proyecto vacío detectado, EN MAESTRA: NO")
        logs.append([row["Cliente"], row["Empresa INDRA/MPS/TCN"], row["N° de Comprobante"], 
                    f"{row['RUC2']}-01-{row['N° de Comprobante']}", "Proyecto vacío", "No procesado"])
        continue
        
    resultadoProyecto = buscar_proyecto(proyecto)
    print("proyecto : " + proyecto + ", resultadoProyecto: " + resultadoProyecto.to_string())
    
    # Actualizar la columna "EN MAESTRA" según el resultado
    if resultadoProyecto is not None and len(resultadoProyecto) > 0:
        df_excel.at[idx, "EN MAESTRA"] = "OK"
        print("Se encontraron resultados : " + proyecto + ", EN MAESTRA: OK")
        print("*************")
        cliente = row["Cliente"]
        empresa = row["Empresa INDRA/MPS/TCN"]
        num_comprobante = row["N° de Comprobante"]
        print(f"num_comprobante: {num_comprobante}")
        print(f"2 num_comprobante: {num_comprobante}")
        codigo_archivo = f"{row['RUC2']}-01-{num_comprobante}"  # Concatenar E y F
        print(f"codigo_archivo: {codigo_archivo}")

        # Crear carpetas
        cliente_dir = os.path.join(BASE_DIR, cliente).replace(".","")
        empresa_dir = os.path.join(cliente_dir, empresa).replace(".","")
        print("cliente_dir : " + cliente_dir)
        print("empresa : " + empresa)
        os.makedirs(empresa_dir, exist_ok=True)
        
        # Transformar número de comprobante
        codigo_busqueda = transformar_codigo(num_comprobante)
        # Verificar estructura y tamaño del CSV antes de buscar
        print(f" Verificando DataFrame antes de buscar:")
        print(f"   Filas en el CSV: {len(df_csv)}")
        print(f"   Columnas en el CSV: {df_csv.shape[1]}")
        print(f"   Columna de búsqueda: {columna_comprobante} de {df_csv.shape[1]-1} (índice 0-based)")

        # Mostrar algunos valores de la columna donde estamos buscando para confirmar formato
        muestra_valores = df_csv.iloc[:5, columna_comprobante].tolist()
        print(f"   Muestra de valores en columna de búsqueda: {muestra_valores}")

        # Buscar en el CSV usando la columna identificada dinámicamente
        print(f" Buscando comprobante con código: '{codigo_busqueda}' en la columna {columna_comprobante}")
        resultado = df_csv[df_csv.iloc[:, columna_comprobante] == codigo_busqueda]
        
        if resultado.empty:
            codigo_busqueda = transformar_codigo_3(num_comprobante)
            codigo_archivo = codigo_archivo.replace("-01-","-03-")
            resultado = df_csv[df_csv.iloc[:, columna_comprobante] == codigo_busqueda]
            
        if not resultado.empty:
            print("resultado: OK " + codigo_busqueda)
            valor_columna_d = resultado.iloc[0, columna_proyecto]
            
            # Crear carpeta del proyecto
            carpeta_proyecto = os.path.join(empresa_dir, valor_columna_d)
            os.makedirs(carpeta_proyecto, exist_ok=True)
            
            # Crear subcarpeta con el número de factura
            carpeta_factura = os.path.join(carpeta_proyecto, num_comprobante)
            os.makedirs(carpeta_factura, exist_ok=True)
            
            print("codigo_archivo : " + codigo_archivo)
            
            # Copiar archivos a la carpeta de la factura en lugar de la carpeta del proyecto
            archivos_copiados = buscar_y_copiar_archivos(codigo_archivo, carpeta_factura)
            
            oc = row["OC-OS"]
            cliente = row["Cliente"]
            nfactura = row["N° de Comprobante"]

            if oc and str(oc).strip():
                print(f"Cliente : {cliente}")
                #Caso La positiva 
                if "POSITIVA" in cliente:
                    print(f"Se busca en la combinacion la positiva")
                    #Combinacion NFactura_OC
                    combinacion_LaPositiva = f"{nfactura}_{oc}"
                    archivos_copiados_oc = buscar_y_copiar_archivos_OC(combinacion_LaPositiva, 
                    PATH_DIRECTORY_FILES, 
                    carpeta_factura)
                else:
                    print(f"Se busca solo con el OC")
                    archivos_copiados_oc = buscar_y_copiar_archivos_OC(oc, 
                    PATH_DIRECTORY_FILES, 
                    carpeta_factura)  # Usar la carpeta de factura como destino
                
                # Combinar los resultados de archivos copiados
                if archivos_copiados_oc:
                    archivos_copiados.extend(archivos_copiados_oc)

            # --- BUSQUEDA POR NR-CR (Añadir después de la búsqueda por OC) ---
            nr = row["Número de Recepción (NR-CR)"]
            if nr and str(nr).strip():
                print(f" Buscando archivos adicionales con el NR: {nr}")
                # Reutilizamos la misma función de búsqueda
                archivos_copiados_nr = buscar_y_copiar_archivos_OC(
                    nr, 
                    PATH_DIRECTORY_FILES, 
                    carpeta_factura
                )
                
                # Si encontró algo, lo agregamos a la lista de archivos copiados para el log
                if archivos_copiados_nr:
                    print(f" Se encontraron y copiaron {len(archivos_copiados_nr)} archivos por NR")
                    archivos_copiados.extend(archivos_copiados_nr)
                else:
                    print(f" No se encontraron archivos para el NR: {nr}")
            
            log_estado = "Completado" if archivos_copiados else "Faltan archivos"
            carpeta_final = carpeta_factura  # Para el log, guardar la ruta completa
        else:
            carpeta_final = "No encontrado"
            log_estado = "No encontrado en CSV"
        
        logs.append([cliente, empresa, num_comprobante, codigo_archivo, carpeta_final, log_estado])
    else:
        df_excel.at[idx, "EN MAESTRA"] = "NO"
        print("No se encontraron resultados : " + proyecto + ", EN MAESTRA: NO")
        logs.append([row["Cliente"], row["Empresa INDRA/MPS/TCN"], row["N° de Comprobante"], 
                     f"{row['RUC2']}-01-{row['N° de Comprobante']}", "No encontrado en maestra", "No procesado"])

# Guardar el log en Excel
df_log = pd.DataFrame(logs, columns=["Cliente", "Empresa", "N° Comprobante", "Archivo Buscado", "Carpeta Destino", "Estado"])
df_log.to_excel(LOG_FILE, index=False)

# Guardar el Excel original con la nueva columna
excel_actualizado = os.path.join(BASE_PATH, "3.file_table_xml.xlsx")
df_excel.to_excel(excel_actualizado, index=False)
print(f"\n Excel actualizado guardado en: {excel_actualizado}")

# NUEVO: Actualizar el Excel histórico con los datos que incluyen la columna "EN MAESTRA"
print("\n Actualizando Excel histórico con los datos procesados...")
historico_actualizado = append_to_existing_excel(df_excel)

if historico_actualizado:
    print(f"\n Excel histórico actualizado correctamente: {historico_actualizado}")
else:
    print("\n No se pudo actualizar el Excel histórico. Revise los logs para más detalles.")

print("\n Proceso terminado. Log guardado en log.xlsx")

# Imprimir un mensaje especial para que el script automatizado pueda capturar el directorio de salida
print(f"OUTPUT_DIRECTORY={BASE_PATH}")