import xml.etree.ElementTree as ET
import re
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os
import shutil
import argparse
import sys
from datetime import datetime
import io

# ==================== CONFIGURACIÓN UTF-8 PARA WINDOWS ====================
# Forzar UTF-8 en stdout y stderr para evitar errores de codificación
if sys.platform == 'win32':
    if sys.stdout.encoding != 'utf-8':
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    if sys.stderr.encoding != 'utf-8':
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
# ==========================================================================


# Configurar la ruta base
# Si hay argumentos, tomar el primero como ruta base, si no, usar la ruta por defecto
if len(sys.argv) > 1:
    BASE_PATH = sys.argv[1]
else:
    BASE_PATH = "inboxFacturas/RV_ Facturación Perú 16012025"
DIRECTORY_FILE_XML = f"{BASE_PATH}/comprobantes_XML"
NEW_EXCEL = Path(DIRECTORY_FILE_XML).parent / "3.file_table_xml.xlsx"  # Ruta para el Excel nuevo
EXISTING_EXCEL = r"D:\facturas_bot\Registros Historico\3.Historicov2.xlsx"
CSV_FILE_TABLE = f"{BASE_PATH}/tabla_1.csv"

df_csv = pd.read_csv(CSV_FILE_TABLE, delimiter='|', dtype=str, header=0, on_bad_lines='warn').fillna("")


def identificar_columnas_csv(df, archivo_log=None):
    """
    Identifica dinámicamente las columnas del CSV basadas en patrones:
    - Columna de comprobante: acepta F001-038923, 01-F001--0389237 y F003--0003600
    - Columna de proyecto: 6 caracteres alfanuméricos
    Retorna: (índice_columna_comprobante, índice_columna_proyecto)
    """
    import re
    from datetime import datetime
    import os

    if archivo_log is None:
        # Usar carpeta LOGS/read_xml (relativa a raíz del proyecto)
        project_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        logs_dir = os.path.join(project_root, "LOGS", "read_xml")
        os.makedirs(logs_dir, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        archivo_log = os.path.join(logs_dir, f"log_identificacion_columnas_{timestamp}.txt")

    def log(mensaje):
        with open(archivo_log, "a", encoding="utf-8") as f:
            f.write(mensaje + "\n")

    with open(archivo_log, "w", encoding="utf-8") as f:
        f.write(f"=== LOG DE IDENTIFICACIÓN DE COLUMNAS ===\n")
        f.write(f"Fecha y hora: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")

    indice_columna_comprobante = None
    indice_columna_proyecto = None

    # PATRÓN FLEXIBLE para ambos formatos de comprobante
    patron_comprobante = re.compile(
        r'^([A-Z]{1}\d{3}-\d{6}|\d{2}-[A-Z0-9]{1,4}--?\d{6,7}|[A-Z]{1}\d{3}--\d{7})$'
    )
    patron_proyecto = re.compile(r'^[A-Za-z0-9]{6}$')

    log(f"Patrones a buscar:")
    log(f"- Comprobante: {patron_comprobante.pattern}")
    log(f"- Proyecto: {patron_proyecto.pattern} (6 caracteres alfanuméricos)")

    # Salta la primera fila de encabezados si la tienes
    df_analisis = df.iloc[1:]

    log(f"\nAnalizando DataFrame con {len(df_analisis)} filas y {len(df.columns)} columnas")
    num_filas_analizar = min(10, len(df_analisis))
    log(f"Se analizarán las primeras {num_filas_analizar} filas")

    for col_idx in range(len(df.columns)):
        nombre_columna = df.columns[col_idx] if hasattr(df, 'columns') and len(df.columns) > col_idx else f'Columna {col_idx}'
        log(f"\n--- Analizando columna {col_idx}: {nombre_columna} ---")
        valores_comprobante = 0
        valores_proyecto = 0

        for idx, row in enumerate(df_analisis.head(num_filas_analizar).itertuples(index=False)):
            valor = str(row[col_idx]).strip()
            log(f"  Fila {idx+1}: Valor = '{valor}'")

            if valor and valor.lower() != 'nan':
                if patron_comprobante.match(valor):
                    valores_comprobante += 1
                    log(f"    OK Coincide con patrón de comprobante")
                else:
                    log(f"    ✗ No coincide con patrón de comprobante")

                if patron_proyecto.match(valor):
                    valores_proyecto += 1
                    log(f"    OK Coincide con patrón de proyecto (6 caracteres alfanuméricos)")
                else:
                    log(f"    ✗ No coincide con patrón de proyecto")
                    if len(valor) != 6:
                        log(f"      -> La longitud es {len(valor)}, se esperaban 6 caracteres")
                    if not re.match(r'^[A-Za-z0-9]*$', valor):
                        log(f"      -> Contiene caracteres que no son alfanuméricos")
            else:
                log(f"    ! Valor vacío o NaN - ignorado")

        log(f"  Resumen columna {col_idx}:")
        log(f"    - Valores que coinciden con patrón de comprobante: {valores_comprobante}/{num_filas_analizar}")
        log(f"    - Valores que coinciden con patrón de proyecto: {valores_proyecto}/{num_filas_analizar}")

        if valores_comprobante > num_filas_analizar / 2 and indice_columna_comprobante is None:
            indice_columna_comprobante = col_idx
            log(f"   Columna de comprobantes identificada: {col_idx} (patrón comprobante flexible)")

        if valores_proyecto > num_filas_analizar / 2 and indice_columna_proyecto is None:
            indice_columna_proyecto = col_idx
            log(f"   Columna de proyecto identificada: {col_idx} (patrón 6 caracteres alfanuméricos)")

    log("\n=== RESULTADOS FINALES ===")
    if indice_columna_comprobante is None:
        indice_columna_comprobante = 0
        log(f"!️ No se pudo identificar automáticamente la columna de comprobantes. Usando predeterminado: 0")
    else:
        log(f" Columna de comprobantes: {indice_columna_comprobante}")

    if indice_columna_proyecto is None:
        indice_columna_proyecto = 3
        log(f"!️ No se pudo identificar automáticamente la columna de proyecto. Usando predeterminado: 3")
    else:
        log(f" Columna de proyecto: {indice_columna_proyecto}")

    log(f"\nLog guardado en: {os.path.abspath(archivo_log)}")
    return indice_columna_comprobante, indice_columna_proyecto


def normaliza_comprobante(comprobante):
    """
    Recibe cualquier comprobante tipo F001-038941, F001--0038941, F001-38941, F001-0038941,
    y lo devuelve siempre como F001-038941 (un guion y 6 dígitos).
    """
    import re
    comprobante = comprobante.replace("--", "-").replace(" ", "").strip()
    match = re.match(r"([A-Z]{1}\d{3})-(\d+)", comprobante)
    if match:
        serie, numero = match.groups()
        return f"{serie}-{int(numero):06d}"  # ceros a la izquierda para 6 dígitos
    return comprobante.strip()

# Identificar las columnas del CSV
columna_comprobante, columna_proyecto = identificar_columnas_csv(df_csv)
print(f"Se usará la columna {columna_comprobante} para buscar comprobantes y la columna {columna_proyecto} para obtener el proyecto")


# 2. Normalizar la columna de comprobantes del CSV
df_csv.iloc[:, columna_comprobante] = df_csv.iloc[:, columna_comprobante].astype(str).apply(normaliza_comprobante)


def variantes_comprobante(comprobante):
    variantes = set()
    # Si ya tiene doble guion, también busca con un solo guion y 6 dígitos
    if "--" in comprobante:
        base = comprobante.replace("--", "-")
        if base.count("-") == 1:
            serie, numero = base.split("-")
            variantes.add(f"{serie}-{numero.lstrip('0')}")
            variantes.add(f"{serie}-{numero.zfill(6)}")
        variantes.add(comprobante)
    # Si tiene un solo guion, genera la versión con doble guion y 7 dígitos
    elif comprobante.count("-") == 1:
        serie, numero = comprobante.split("-")
        variantes.add(comprobante)
        variantes.add(f"{serie}--{numero.zfill(7)}")
    else:
        variantes.add(comprobante)
    return list(variantes)

import re
def buscar_proyecto_tabla_CSV(df_csv, codigo_busqueda, columna_comprobante, columna_proyecto):
    import re
    variantes = set()
    
    # Variante exacta como viene del XML
    variantes.add(codigo_busqueda)
    
    # Normaliza a formato F001-0039417 (con seis o siete dígitos)
    match = re.match(r'([A-Z]{1}\d{3})-(\d+)', codigo_busqueda)
    if match:
        serie, numero = match.groups()
        num_int = int(numero)
        
        # Generar múltiples formatos de número
        num_6d = f"{num_int:06d}"  # 6 dígitos: 041890
        num_7d = f"{num_int:07d}"  # 7 dígitos: 0041890
        
        # VARIANTES SIN PREFIJO
        variantes.add(f"{serie}-{num_6d}")
        variantes.add(f"{serie}-{num_7d}")
        variantes.add(f"{serie}--{num_6d}")
        variantes.add(f"{serie}--{num_7d}")
        
        # VARIANTES CON PREFIJO "01-"
        variantes.add(f"01-{serie}-{num_6d}")
        variantes.add(f"01-{serie}-{num_7d}")
        variantes.add(f"01-{serie}--{num_6d}")
        variantes.add(f"01-{serie}--{num_7d}")
        
        # VARIANTES CON PREFIJO "03-"
        variantes.add(f"03-{serie}-{num_6d}")
        variantes.add(f"03-{serie}-{num_7d}")
        variantes.add(f"03-{serie}--{num_6d}")
        variantes.add(f"03-{serie}--{num_7d}")
    
    print(f" Buscando comprobante: '{codigo_busqueda}'")
    print(f" Variantes generadas ({len(variantes)}): {sorted(variantes)}")
    
    # !️ CAMBIO CRÍTICO: Usar la columna correcta identificada dinámicamente
    columna_texto = df_csv.iloc[:, columna_comprobante].astype(str).str.strip()
    
    print(f" Primeros 5 valores de la columna {columna_comprobante} del CSV:")
    print(columna_texto.head(5).tolist())
    
    # Búsqueda exacta primero
    for variante in sorted(variantes):
        resultado = df_csv[columna_texto == variante]
        if not resultado.empty:
            proyecto = resultado.iloc[0, columna_proyecto]
            print(f" ¡ENCONTRADO! Variante exacta: '{variante}' -> Proyecto: '{proyecto}'")
            return proyecto
    
    # Búsqueda con "contiene" como fallback
    for variante in sorted(variantes):
        resultado = df_csv[columna_texto.str.contains(variante, regex=False, na=False)]
        if not resultado.empty:
            proyecto = resultado.iloc[0, columna_proyecto]
            print(f" ¡ENCONTRADO! (contiene) Variante: '{variante}' -> Proyecto: '{proyecto}'")
            return proyecto
    
    print(f"ERROR No se encontró ninguna variante del comprobante '{codigo_busqueda}' en el CSV")
    return None
def transformar_codigo(comprobante):
    partes = comprobante.split("-")
    if len(partes) == 2:
        serie, numero = partes
        numero = numero.zfill(7)
        return f"{serie}--{numero}"
    return comprobante



def extract_invoice_data(xml_content):
    # namespaces del XML
    namespaces = {
        'cbc': 'urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2',
        'cac': 'urn:oasis:names:specification:ubl:schema:xsd:CommonAggregateComponents-2',
        'ext': 'urn:oasis:names:specification:ubl:schema:xsd:CommonExtensionComponents-2',
        'ds': 'http://www.w3.org/2000/09/xmldsig#',
        'default': 'urn:oasis:names:specification:ubl:schema:xsd:Invoice-2'
    }
    
    try:
        if xml_content.startswith('<?xml'):
            xml_content = re.sub(r'<\?xml[^>]+\?>', '', xml_content)
        
        root = ET.fromstring(xml_content)
        
        # Creamos un diccionario para almacenar los datos extraídos
        extracted_data = {}
        
        # 1. Cliente (Señores)
        customer_party = root.find('.//cac:AccountingCustomerParty/cac:Party', namespaces)
        if customer_party is not None:
            customer_name = customer_party.find('.//cac:PartyLegalEntity/cbc:RegistrationName', namespaces)
            if customer_name is not None:
                extracted_data['Cliente'] = customer_name.text
            else:
                extracted_data['Cliente'] = "No encontrado"
        
        # 2. RUC Cliente
        customer_id = root.find('.//cac:AccountingCustomerParty/cac:Party/cac:PartyIdentification/cbc:ID', namespaces)
        if customer_id is not None:
            extracted_data['RUC'] = customer_id.text
        
        # 3. Proyecto (Buscamos en AdditionalDocumentReference o en la descripción)
        # Coloco mas ABAJO
        
        # 4. Empresa (Supplier)
        supplier_name = root.find('.//cac:AccountingSupplierParty/cac:Party/cac:PartyLegalEntity/cbc:RegistrationName', namespaces)
        if supplier_name is not None:
                extracted_data['Empresa INDRA/MPS/TCN'] = supplier_name.text
        
        # 5. RUC Proveedor
        supplier_id = root.find('.//cac:AccountingSupplierParty/cac:Party/cac:PartyIdentification/cbc:ID', namespaces)
        if supplier_id is not None:
            extracted_data['RUC2'] = supplier_id.text

        # 6. Número de Factura
        invoice_id = root.find('./cbc:ID', namespaces)
        if invoice_id is not None:
            print(invoice_id.text)
            extracted_data['N° de Comprobante'] = invoice_id.text
            
        
        # 6.1. Proyecto
        if invoice_id is not None:
            invoiceNumber = invoice_id.text
            invoiceNumber = normaliza_comprobante(invoiceNumber)
            #invoiceNumber = transformar_codigo(invoiceNumber)
            print("InvoiceNumber : " + invoiceNumber)
            print("\n====== DEBUG BUSQUEDA DE PROYECTO ======")
            print(f"Comprobante extraído del XML: '{invoiceNumber}'")
            print("Primeros comprobantes en CSV (col 0):")
            print(df_csv.iloc[:, columna_comprobante].head(10).tolist())
            print("Primeros proyectos en CSV (col 3):")
            print(df_csv.iloc[:, columna_proyecto].head(10).tolist())
            print("========================================\n")
            #project = buscar_proyecto_tabla_CSV(df_csv,invoiceNumber,columna_proyecto)
            project = buscar_proyecto_tabla_CSV(df_csv, invoiceNumber, columna_comprobante, columna_proyecto)

            if project is not None:
                print("project : " + str(project))  # Convertir el valor en string si no es None
            else:
                print("No se encontró la factura : " + invoiceNumber +", por lo tanto no tiene proyecto.")
            extracted_data['Proyecto'] = project

        # 7. Fecha de Emisión
        issue_date = root.find('./cbc:IssueDate', namespaces)
        if issue_date is not None:
            extracted_data['Fecha de Envío'] = issue_date.text
        
        # 8. Divisa
        note_text = ""
        note_element = root.find('./cbc:Note[@languageLocaleID="1000"]', namespaces)
        if note_element is not None and note_element.text:
            note_text = note_element.text.strip()
        extracted_data['Divisa'] = note_text

        
        # 9. Tipo de Impuesto
        tax_type = ""
        tax_subtotal = root.find('.//cac:TaxSubtotal/cac:TaxCategory/cac:TaxScheme/cbc:Name', namespaces)
        if tax_subtotal is not None:
            tax_type = tax_subtotal.text
        extracted_data['Tipo de Impuesto'] = tax_type
        

        # 10. Condiciones de pago - Días
        payment_days = root.find('.//cbc:Note[@languageID="L"]', namespaces)
        if payment_days is not None and "DIAS" in payment_days.text.upper():
            extracted_data['Condición de pago'] = payment_days.text
        else:
            # Alternativa: buscar en PaymentTerms
            payment_terms = root.find('.//cac:PaymentTerms/cbc:Note', namespaces)
            if payment_terms is not None and "DIAS" in payment_terms.text.upper():
                extracted_data['Condición de pago'] = payment_terms.text
        
        # 11. Valor venta (monto sin IGV)
        monetary_total = root.find('.//cac:LegalMonetaryTotal', namespaces)
        if monetary_total is not None:
            line_extension = monetary_total.find('./cbc:LineExtensionAmount', namespaces)
            if line_extension is not None:
                extracted_data['Valor Venta'] = line_extension.text
        
        # 12. IGV (18%)
        tax_total = root.find('.//cac:TaxTotal/cbc:TaxAmount', namespaces)
        if tax_total is not None:
            extracted_data['IGV (18%)'] = tax_total.text
        
        # 13. Total a pagar
        if monetary_total is not None:
            payable_amount = monetary_total.find('./cbc:PayableAmount', namespaces)
            if payable_amount is not None:
                extracted_data['TOTAL'] = payable_amount.text
        
        # 14. Número de Recepción (NR-CR)
        nr = ""
        description = root.find('.//cac:Item/cbc:Description', namespaces)
        if description is not None and description.text:
            # Reemplazar codificaciones URL comunes
            texto_limpio = description.text.replace('%5D', ']').replace('%5B', '[')
            
            # Patrón flexible para NR con múltiples formatos posibles
            nr_match = re.search(r'(?:NR[:#._\s]*|N[°º]?\s*NR[:#._\s]*|Nro\.?\s*NR[:#._\s]*|Numero\s*(?:de)?\s*Recepcion\s*(?:N[°º]?)?\s*)(\d{1,15}(?:-\d{1,10})?)', texto_limpio)
            if nr_match:
                nr = nr_match.group(1)
            else:
                # Buscar alternativas como texto antes o después de una barra
                nr_match = re.search(r'[/|]\s*NR\s*[:|=]?\s*(\d{1,15}(?:-\d{1,10})?)', texto_limpio)
                if nr_match:
                    nr = nr_match.group(1)

        # Si no se encontró en description, buscar en OrderReference
        if not nr:
            order_reference = root.find('.//cac:OrderReference/cbc:ID', namespaces)
            if order_reference is not None and order_reference.text:
                texto_limpio = order_reference.text.replace('%5D', ']').replace('%5B', '[')
                nr_match = re.search(r'(?:NR[:\s]\s*|N[°º]?\s*NR[:\s]\s*|Nro\.?\s*NR[:\s]\s*|Numero\s*(?:de)?\s*Recepcion\s*(?:N[°º]?)?\s*)(\d{1,15}(?:-\d{1,10})?)', texto_limpio)
                if nr_match:
                    nr = nr_match.group(1)
                else:
                    nr_match = re.search(r'[/|]\s*NR\s*[:|=]?\s*(\d{1,15}(?:-\d{1,10})?)', texto_limpio)
                    if nr_match:
                        nr = nr_match.group(1)

        # Buscar también en Note si todavía no se encontró
        if not nr:
            notes = root.findall('.//cbc:Note', namespaces)
            for note in notes:
                if note is not None and note.text:
                    texto_limpio = note.text.replace('%5D', ']').replace('%5B', '[')
                    nr_match = re.search(r'(?:NR[:\s]\s*|N[°º]?\s*NR[:\s]\s*|Nro\.?\s*NR[:\s]\s*|Numero\s*(?:de)?\s*Recepcion\s*(?:N[°º]?)?\s*)(\d{1,15}(?:-\d{1,10})?)', texto_limpio)
                    if nr_match:
                        nr = nr_match.group(1)
                        break
                    else:
                        nr_match = re.search(r'[/|]\s*NR\s*[:|=]?\s*(\d{1,15}(?:-\d{1,10})?)', texto_limpio)
                        if nr_match:
                            nr = nr_match.group(1)
                            break

        extracted_data['Número de Recepción (NR-CR)'] = nr
        
        # 15. OC
        oc = ""
        descriptions = root.findall('.//cac:Item/cbc:Description', namespaces)
        for desc in descriptions:
            print(desc)
            if desc is not None and desc.text:
                texto_limpio = desc.text.replace('%5D', ']').replace('%5B', '[')
                print(texto_limpio)
                oc_match = re.search(r'(?:OC\s+|Orden de compra\s+(?:N[°º]?\s*)?)(\w{3,12}(?:-\w{4})?)', texto_limpio)
                print("resultado")
                
                if oc_match:
                    oc = oc_match.group(1)
                    print(oc)
                    extracted_data['OC-OS'] = oc
                    break
        
        # 16. Descripción (Primera Fila)
        first_line_description = ""
        invoice_lines = root.findall('.//cac:InvoiceLine', namespaces)
        if invoice_lines and len(invoice_lines) > 0:
            first_item_desc = invoice_lines[0].find('.//cac:Item/cbc:Description', namespaces)
            if first_item_desc is not None and first_item_desc.text:
                first_line_description = first_item_desc.text
        extracted_data['Descripción (Primera Fila)'] = os.path.basename(BASE_PATH) + "-" +first_line_description
        
        extracted_data['ENVAR CORREO'] = "NO"
        extracted_data['ESTADO'] = "SIN PROCESAR"
        
        return extracted_data
    
    except Exception as e:
        print(f"Error al procesar XML: {str(e)}")
        return {"error": str(e)}

def create_excel_with_headers(output_file):
    """
    Crea un archivo Excel con los encabezados especificados y formato
    """
    headers = [
        "Cliente", "RUC", "Proyecto", "Empresa INDRA/MPS/TCN", "RUC2", 
        "N° de Comprobante", "Fecha de Envío", "Divisa", "Tipo de Impuesto",
        "Condición de pago", "Valor Venta", "IGV (18%)", "TOTAL", 
        "OC-OS", "Número de Recepción (NR-CR)", "Descripción (Primera Fila)",
        "ENVAR CORREO",
        "ESTADO"
    ]
    
    # Crear DataFrame vacío con las columnas
    df = pd.DataFrame(columns=headers)
    
    # Guardar a Excel
    df.to_excel(output_file, index=False, sheet_name='Facturas')
    
    # Dar formato con openpyxl
    wb = openpyxl.load_workbook(output_file)
    ws = wb['Facturas']
    
    # Estilo para encabezados
    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    value_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    description_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # Aplicar bordes a todas las celdas
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Formato para encabezados
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        
        # Aplicar color según tipo de columna
        if header in ["Valor Venta", "IGV (18%)", "TOTAL","ENVAR CORREO","ESTADO"]:
            cell.fill = value_fill
        elif header in ["Descripción (Primera Fila)"]:
            cell.fill = description_fill
        else:
            cell.fill = header_fill
    
    # Ajustar ancho de columnas
    for col_num, _ in enumerate(headers, 1):
        col_letter = openpyxl.utils.get_column_letter(col_num)
        if col_num in [1, 3, 4, 16]:  # Nombres, proyectos y descripción más anchos
            ws.column_dimensions[col_letter].width = 25
        elif col_num in [6, 7, 9, 10]:  # Fechas, condiciones con ancho medio
            ws.column_dimensions[col_letter].width = 15
        else:
            ws.column_dimensions[col_letter].width = 12
    
    wb.save(output_file)
    return output_file

def process_xml_files():
    """
    Procesa todos los archivos XML en el directorio y:
    1. Genera un nuevo Excel con los datos
    2. Agrega los mismos datos a un Excel existente
    """
    # Verificar si el directorio de XMLs existe
    directory = Path(DIRECTORY_FILE_XML)
    if not directory.exists():
        print(f"El directorio {DIRECTORY_FILE_XML} no existe. Creando directorio...")
        directory.mkdir(parents=True, exist_ok=True)
        print(f"Directorio creado. Por favor, coloque los archivos XML en {DIRECTORY_FILE_XML}")
        return None, None
    
    # Buscar archivos XML en el directorio
    xml_files = list(directory.glob("*.xml"))
    if not xml_files:
        print(f"No se encontraron archivos XML en {DIRECTORY_FILE_XML}")
        return None, None
    
    print(f"Procesando {len(xml_files)} archivos XML...")
    
    # Lista para almacenar los datos de todas las facturas
    all_invoice_data = []
    
    # Procesar cada archivo XML y extraer datos
    for file in xml_files:
        try:
            with open(file, 'r', encoding='utf-8') as f:
                xml_content = f.read()
                print(f"\nProcesando: {file.name}")
            
            # Extraer datos del XML
            invoice_data = extract_invoice_data(xml_content)
            
            if 'error' in invoice_data:
                print(f"Error al procesar {file.name}: {invoice_data['error']}")
                continue
            
            # Agregar los datos a la lista
            all_invoice_data.append(invoice_data)
            
        except Exception as e:
            print(f"Error al procesar el archivo {file.name}: {str(e)}")
    
    # Si no hay datos procesados, salir
    if not all_invoice_data:
        print("No se pudieron procesar datos de ningún XML.")
        return None, None
    
    # 1. Crear un nuevo Excel con todos los datos
    new_excel_path = create_and_populate_new_excel(all_invoice_data)
    
    # 2. Comentado: Agregar los mismos datos al Excel existente
    # existing_excel_path = append_to_existing_excel(all_invoice_data)
    existing_excel_path = None  # No se agrega al Excel histórico
    
    return new_excel_path, existing_excel_path
    #return new_excel_path

def create_and_populate_new_excel(invoice_data_list):
    """
    Crea un nuevo archivo Excel y lo llena con los datos de las facturas
    """
    try:
        # Crear archivo Excel con encabezados y formato
        output_file = create_excel_with_headers(NEW_EXCEL)
        
        # Cargar el workbook para añadir datos
        wb = openpyxl.load_workbook(output_file)
        ws = wb['Facturas']
        
        # Agregar datos a partir de la fila 2
        row_num = 2
        
        # Obtener encabezados
        headers = [cell.value for cell in ws[1]]
        
        # Agregar cada factura al Excel
        for invoice_data in invoice_data_list:
            for col_num, header in enumerate(headers, 1):
                if header in invoice_data:
                    ws.cell(row=row_num, column=col_num, value=invoice_data[header])
            
            row_num += 1
        
        # Guardar el archivo Excel
        wb.save(output_file)
        print(f"\nSe ha creado exitosamente el nuevo archivo Excel: {output_file}")
        return output_file
        
    except Exception as e:
        print(f"Error al crear nuevo Excel: {str(e)}")
        return None

"""
# Función comentada para que no escriba en el Excel histórico
def append_to_existing_excel(invoice_data_list):
    # Agrega datos de facturas a un archivo Excel existente
    
    existing_excel = EXISTING_EXCEL

    try:
        # Verificar si el archivo Excel existe
        if not Path(existing_excel).exists():
            print(f"El archivo Excel {existing_excel} no existe. Creando una copia del nuevo Excel...")

            # Si el archivo nuevo existe, hacer una copia
            if Path(NEW_EXCEL).exists():
                shutil.copy2(NEW_EXCEL, existing_excel)
                print(f"Se ha creado una copia del nuevo Excel en: {existing_excel}")
                return existing_excel
            else:
                # Si no existe el nuevo, crear uno desde cero
                output_file = create_excel_with_headers(existing_excel)

                # Ahora agregar los datos al Excel recién creado
                wb = openpyxl.load_workbook(output_file)
                ws = wb.active

                # Obtener encabezados
                headers = [cell.value for cell in ws[1]]

                # Empezar desde la fila 2
                row_num = 2

                # Agregar cada factura al Excel
                for invoice_data in invoice_data_list:
                    for col_num, header in enumerate(headers, 1):
                        if header in invoice_data:
                            ws.cell(row=row_num, column=col_num, value=invoice_data[header])

                    row_num += 1

                # Guardar el archivo Excel
                wb.save(output_file)
                print(f"\nSe ha creado y poblado el archivo Excel: {output_file}")
                return output_file

        # Si el archivo existe, cargar datos y verificar duplicados
        df_excel = pd.read_excel(existing_excel)

        # Cargar el Excel existente con openpyxl
        wb = openpyxl.load_workbook(existing_excel)
        ws = wb.active  # Usar la hoja activa

        # Obtener la siguiente fila para agregar datos
        row_num = ws.max_row + 1
        print(f"Agregando datos a partir de la fila {row_num}")

        # Obtener encabezados
        headers = [cell.value for cell in ws[1]]

        # Contador de facturas añadidas
        added_count = 0

        # Agregar cada factura al Excel si no existe ya
        for invoice_data in invoice_data_list:
            # Verificar si la factura ya existe (por número de comprobante)
            comprobante = invoice_data.get('N° de Comprobante', '')
            if comprobante and 'N° de Comprobante' in df_excel.columns and comprobante in df_excel['N° de Comprobante'].values:
                print(f"El comprobante {comprobante} ya existe en el Excel existente. Omitiendo...")
                continue

            # Agregar la factura al Excel
            for col_num, header in enumerate(headers, 1):
                if header in invoice_data:
                    ws.cell(row=row_num, column=col_num, value=invoice_data[header])

            row_num += 1
            added_count += 1

        # Guardar el archivo Excel
        wb.save(existing_excel)
        print(f"\nSe han agregado {added_count} nuevas facturas al archivo Excel existente: {existing_excel}")
        return existing_excel

    except Exception as e:
        print(f"Error al agregar datos al Excel existente: {str(e)}")
        return None
"""
    
if __name__ == "__main__":
    # Mostrar la ruta base que se está utilizando
    print(f"Utilizando ruta base: {BASE_PATH}")
    
    # Procesar los archivos XML para crear un nuevo Excel y agregar datos al existente
    new_excel, existing_excel = process_xml_files()
    #new_excel = process_xml_files()

    if new_excel:
        print(f"Archivo Excel nuevo creado: {new_excel}")
    else:
        print("No se pudo crear el archivo Excel nuevo.")
    
    if existing_excel:
        print(f"Datos agregados al archivo Excel existente: {existing_excel}")
    else:
        print("No se pudieron agregar datos al archivo Excel existente.")