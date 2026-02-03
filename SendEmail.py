import pandas as pd
import openpyxl
import smtplib
import logging
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from pathlib import Path
import os
import mimetypes
from email.mime.base import MIMEBase
from email import encoders

# Configuración de rutas
EXCEL_FILE = r"D:\facturas_bot\Registros Historico\3.Historicov2.xlsx"  # Archivo de facturas
TEMPLATES_FILE = r"D:\facturas_bot\Maestra\Robot 2_Estructura Carpetas Factura SSFF 03.03.2025.xlsx"  # Archivo con plantillas de correo por cliente
TEMPLATES_SHEET = "Configuracion_Correos"  # Nombre de la hoja con las plantillas
LOG_FILE = "monitor_excel.log" 

# Credenciales y configuración
EMAIL = "alertasflm@indra.es"
PASSWORD = "es8EaB63"

# Configuración de servidor SMTP
SMTP_SERVER = "smtp.indra.es"
SMTP_PORT = 587
USE_TLS = True

# Configuración de logging
logging.basicConfig(
    filename=LOG_FILE,
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger()
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO)
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
console_handler.setFormatter(formatter)
logger.addHandler(console_handler)

def check_excel_file_exists(file_path):
    """Verifica si un archivo Excel existe"""
    if not Path(file_path).exists():
        logger.error(f"El archivo Excel {file_path} no existe")
        return False
    return True

def read_excel_data(file_path, sheet_name=0):
    """
    Lee los datos de un archivo Excel
    
    Args:
        file_path: Ruta al archivo Excel
        sheet_name: Nombre u índice de la hoja (0 por defecto para la primera hoja)
    
    Returns:
        DataFrame con los datos o None si hay error
    """
    try:
        # Leer el Excel tratando explícitamente las columnas numéricas como texto
        # para evitar la conversión automática a float
        df = pd.read_excel(file_path, sheet_name=sheet_name, dtype={'RUC': str})
        return df
    except Exception as e:
        logger.error(f"Error al leer el archivo Excel {file_path} (hoja: {sheet_name}): {str(e)}")
        return None

def update_excel_status(index, new_status):
    """Actualiza el estado de una fila en el Excel de facturas"""
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        
        headers = [cell.value for cell in ws[1]]
        estado_col_index = headers.index("ESTADO") + 1
        
        excel_row = index + 2
        ws.cell(row=excel_row, column=estado_col_index).value = new_status
        
        wb.save(EXCEL_FILE)
        logger.info(f"Estado actualizado para la fila {index} a '{new_status}'")
        return True
    except Exception as e:
        logger.error(f"Error al actualizar el estado en Excel: {str(e)}")
        return False

def load_email_templates():
    """Carga las plantillas de correo desde el archivo de plantillas"""
    if not check_excel_file_exists(TEMPLATES_FILE):
        logger.error(f"No se pudo encontrar el archivo de plantillas: {TEMPLATES_FILE}")
        return {}
    
    # Verificar primero que la hoja existe en el archivo
    try:
        # Cargar el workbook para ver las hojas disponibles
        wb = openpyxl.load_workbook(TEMPLATES_FILE, read_only=True)
        available_sheets = wb.sheetnames
        
        if TEMPLATES_SHEET not in available_sheets:
            logger.error(f"La hoja '{TEMPLATES_SHEET}' no existe en el archivo. Hojas disponibles: {', '.join(available_sheets)}")
            logger.info(f"Intentando usar la primera hoja: {available_sheets[0]}")
            sheet_to_use = available_sheets[0]
        else:
            sheet_to_use = TEMPLATES_SHEET
    except Exception as e:
        logger.error(f"Error al verificar hojas del archivo: {str(e)}")
        logger.info("Intentando usar la primera hoja por defecto")
        sheet_to_use = 0
    
    # Ahora leer los datos de la hoja
    templates_df = read_excel_data(TEMPLATES_FILE, sheet_to_use)
    if templates_df is None:
        return {}
    
    # Verificar si el archivo tiene las columnas necesarias
    required_columns = ["Cliente", "RUC", "DESTINATARIO TEST", "ASUNTO", "Cuerpo"]
    missing_columns = [col for col in required_columns if col not in templates_df.columns]
    
    if missing_columns:
        logger.error(f"El archivo de plantillas no tiene las columnas requeridas: {', '.join(missing_columns)}")
        logger.info(f"Columnas disponibles: {', '.join(templates_df.columns)}")
        return {}
    
    # Crear un diccionario de plantillas indexado por RUC
    templates = {}
    for _, row in templates_df.iterrows():
        ruc = row.get("RUC")
        client = row.get("Cliente")
        
        if pd.notna(ruc) and pd.notna(row.get("DESTINATARIO TEST")):
            # Mantener el RUC original como clave pero guardar también el RUC limpio
            ruc_str = str(ruc).strip()
            # Almacenar el RUC original para mantener la trazabilidad
            templates[ruc_str] = {
                "cliente": client,
                "destinatario": row["DESTINATARIO TEST"],
                "asunto": row["ASUNTO"] if pd.notna(row["ASUNTO"]) else f"Notificación de Factura",
                "cuerpo": row["Cuerpo"] if pd.notna(row["Cuerpo"]) else None
            }
    
    logger.info(f"Se cargaron {len(templates)} plantillas de correo de la hoja '{sheet_to_use}'")
    return templates

def find_template_by_ruc(templates, ruc):
    """
    Busca la plantilla correspondiente a un RUC
    
    Args:
        templates: Diccionario de plantillas
        ruc: Número de RUC
        
    Returns:
        dict: Plantilla para el RUC o None si no se encuentra
    """
    if not ruc or pd.isna(ruc):
        logger.error("No se proporcionó un RUC válido")
        return None
    
    # Limpiar el RUC eliminando comas, puntos, espacios y otros caracteres no numéricos
    ruc_str = ''.join(c for c in str(ruc) if c.isdigit())
    
    logger.info(f"Buscando plantilla para RUC: '{ruc_str}' (original: '{ruc}')")
    
    # Buscar coincidencia exacta por RUC
    template_index = 0
    for template_ruc, template in templates.items():
        template_index += 1
        # Limpiar también los RUCs del diccionario de plantillas
        # Eliminar .0 al final que puede aparecer cuando Excel interpreta como float
        clean_template_ruc = ''.join(c for c in str(template_ruc) if c.isdigit())
        
        logger.debug(f"Comparando con plantilla #{template_index}: '{clean_template_ruc}' (original: '{template_ruc}')")
        
        if ruc_str == clean_template_ruc:
            logger.info(f"✓ ENCONTRADO: Plantilla #{template_index} para RUC: '{ruc_str}' (coincide con '{template_ruc}')")
            logger.info(f"  Cliente: {template['cliente']}")
            logger.info(f"  Destinatario: {template['destinatario']}")
            logger.info(f"  Cuerpo : {template['cuerpo']}")
            return template
    
    # Mostrar los RUCs disponibles en las plantillas para depuración
    available_rucs = list(templates.keys())
    if available_rucs:
        logger.info(f"Se verificaron {len(available_rucs)} plantillas de RUC pero no se encontró coincidencia")
        # Mostrar una muestra de los RUCs disponibles
        sample_size = min(5, len(available_rucs))
        sample_rucs = available_rucs[:sample_size]
        clean_sample_rucs = [''.join(c for c in str(r) if c.isdigit()) for r in sample_rucs]
        
        logger.info(f"Muestra de RUCs disponibles (originales): {sample_rucs}")
        logger.info(f"Muestra de RUCs disponibles (limpios): {clean_sample_rucs}")
        logger.info(f"RUC buscado (limpio): '{ruc_str}'")
    
    logger.warning(f"❌ NO ENCONTRADO: No hay plantilla para RUC: {ruc_str}")
    return None

def send_email(row_data, templates, attachment_dir=None):
    """
    Envía un correo usando SMTP con contenido personalizado y adjunta todos los archivos
    que se encuentren en el directorio especificado
    
    Args:
        row_data: Datos de la fila actual
        templates: Diccionario de plantillas de correo
        attachment_dir: Ruta al directorio que contiene los archivos a adjuntar (opcional)
    
    Returns:
        bool: True si el correo se envió correctamente, False en caso contrario
    """
    try:
        # Obtener el RUC del cliente
        client_ruc = row_data.get('RUC')
        client_name = row_data.get('Cliente')
        
        if not client_ruc or pd.isna(client_ruc):
            logger.error("No se encontró el RUC del cliente en la fila")
            return False
        
        # Información de log para depuración
        log_client_info = f"{client_name} con ruc {client_ruc}" if pd.notna(client_name) else f"RUC {client_ruc}"
        
        # Buscar la plantilla para este RUC
        template = find_template_by_ruc(templates, client_ruc)
        
        # Si no hay plantilla, no enviar correo
        if not template:
            logger.warning(f"No se encontró plantilla para {log_client_info}, no se enviará correo")
            return False
        
        # Procesar destinatarios (pueden ser múltiples separados por punto y coma)
        destinatarios_raw = template["destinatario"]
        destinatarios_list = [email.strip() for email in destinatarios_raw.split(';') if email.strip()]
        
        if not destinatarios_list:
            logger.error(f"No se encontraron destinatarios válidos para {log_client_info}")
            return False
            
        destinatarios_str = "; ".join(destinatarios_list)
        logger.info(f"Enviando correo a {len(destinatarios_list)} destinatarios: {destinatarios_str}")
        
        asunto_original = template["asunto"]
        cuerpo_personalizado = template["cuerpo"]
        
        # Crear un diccionario de mapeo para variables comunes
        variable_mapping = {
            "{Nro_Factura}": row_data.get("N° de Comprobante", ""),
            "{Orden_Compra}": row_data.get("OC-OS", ""),
            "{TotalMonto}": row_data.get("TOTAL", ""),
            # Agregar más mapeos según sea necesario
        }
        
        # Reemplazar variables en el asunto
        asunto = asunto_original
        
        # Primero aplicar el mapeo de variables personalizadas
        for placeholder, value in variable_mapping.items():
            if pd.notna(value):
                asunto = asunto.replace(placeholder, str(value))
        
        # Luego aplicar el reemplazo estándar por nombre de columna
        for key, value in row_data.items():
            if pd.notna(value):
                placeholder = f"{{{key}}}"
                asunto = asunto.replace(placeholder, str(value))
        
        # Crear el mensaje
        msg = MIMEMultipart()
        msg['From'] = EMAIL
        msg['To'] = destinatarios_str  # Lista de destinatarios separados por punto y coma
        msg['Subject'] = asunto
        
        # Si hay un cuerpo personalizado, usarlo; de lo contrario, usar plantilla HTML
        if cuerpo_personalizado:
            # Reemplazar variables en el cuerpo personalizado
            cuerpo_final = cuerpo_personalizado
            
            # Primero aplicar el mapeo de variables personalizadas
            for placeholder, value in variable_mapping.items():
                if pd.notna(value):
                    cuerpo_final = cuerpo_final.replace(placeholder, str(value))
            
            # Luego aplicar el reemplazo estándar por nombre de columna
            for key, value in row_data.items():
                if pd.notna(value):
                    placeholder = f"{{{key}}}"
                    cuerpo_final = cuerpo_final.replace(placeholder, str(value))
            
            msg.attach(MIMEText(cuerpo_final, 'html'))
        else:
            # Usar la plantilla HTML estándar
            html_content = create_html_content(row_data)
            msg.attach(MIMEText(html_content, 'html'))
        
        # Adjuntar archivos del directorio especificado
        attachment_count = 0
        # Ya no necesitamos quitar puntos aquí porque se construye el path sin puntos desde el inicio
        if attachment_dir and os.path.isdir(attachment_dir):
            for filename in os.listdir(attachment_dir):
                if filename.startswith('.'):
                    continue
                file_path = os.path.join(attachment_dir, filename)
                
                # Verificar que sea un archivo (no un directorio)
                if os.path.isfile(file_path):
                    try:
                        # Determinar el tipo MIME basado en la extensión del archivo
                        content_type, encoding = mimetypes.guess_type(file_path)
                        if content_type is None:
                            # Si no se puede determinar el tipo, usar application/octet-stream
                            content_type = 'application/octet-stream'
                        
                        main_type, sub_type = content_type.split('/', 1)
                        
                        # Abrir y adjuntar el archivo según su tipo
                        with open(file_path, 'rb') as file:
                            attachment = MIMEBase(main_type, sub_type)
                            attachment.set_payload(file.read())
                        
                        # Codificar en base64
                        encoders.encode_base64(attachment)
                        
                        # Agregar encabezado con nombre de archivo
                        attachment.add_header('Content-Disposition', 'attachment', filename=filename)
                        msg.attach(attachment)
                        
                        attachment_count += 1
                        logger.info(f"Archivo adjunto agregado: {filename}")
                    except Exception as e:
                        logger.error(f"Error al adjuntar archivo {file_path}: {str(e)}")
                        # Continuar con el resto de adjuntos aunque uno falle
            
            logger.info(f"Total de archivos adjuntos: {attachment_count}")
        elif attachment_dir:
            logger.warning(f"El directorio de adjuntos no existe o no es accesible: {attachment_dir}")
        
        # Enviar el correo
        server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT, timeout=30)
        
        if USE_TLS:
            server.starttls()
        
        server.login(EMAIL, PASSWORD)
        
        # Enviar el correo a todos los destinatarios
        server.sendmail(EMAIL, destinatarios_list, msg.as_string())
        server.quit()
        
        logger.info(f"Correo enviado exitosamente a {len(destinatarios_list)} destinatarios con {attachment_count} archivos adjuntos")
        return True
        
    except Exception as e:
        logger.error(f"Error al enviar correo: {str(e)}")
        return False
         
def create_html_content(row_data):
    """Crea el contenido HTML del correo basado en los datos de la fila"""
    html = f"""
    <html>
    <head>
        <style>
            table {{
                border-collapse: collapse;
                width: 100%;
            }}
            th, td {{
                border: 1px solid #dddddd;
                text-align: left;
                padding: 8px;
            }}
            th {{
                background-color: #f2f2f2;
            }}
            tr:nth-child(even) {{
                background-color: #f9f9f9;
            }}
        </style>
    </head>
    <body>
        <h2>Notificación de Procesamiento de Factura</h2>
        <p>Se ha marcado una factura para procesamiento con los siguientes datos:</p>
        
        <table>
            <tr>
                <th>Campo</th>
                <th>Valor</th>
            </tr>
    """
    
    # Agregar cada campo al correo
    for column, value in row_data.items():
        if pd.notna(value):  # Solo incluir valores no nulos
            html += f"""
            <tr>
                <td>{column}</td>
                <td>{value}</td>
            </tr>
            """
    
    html += """
        </table>
        <p>Este es un correo automático, por favor no responda a este mensaje.</p>
    </body>
    </html>
    """
    
    return html

def process_pending_emails():
    """Procesa las facturas pendientes para envío de correo"""
    logger.info("Iniciando proceso de revisión del Excel")
    
    # Verificar si el archivo de facturas existe
    if not check_excel_file_exists(EXCEL_FILE):
        return
    
    # Leer datos del Excel de facturas
    df = read_excel_data(EXCEL_FILE)
    if df is None:
        return
    
    # Verificar si las columnas necesarias existen
    required_columns = ["ENVAR CORREO", "ESTADO", "Cliente", "RUC"]
    for col in required_columns:
        if col not in df.columns:
            logger.error(f"La columna '{col}' no existe en el Excel")
            return
    
    # Mostrar las columnas disponibles para referencia (útil para configurar variables)
    logger.info(f"Columnas disponibles en el Excel: {list(df.columns)}")
    
    # Cargar las plantillas de correo
    templates = load_email_templates()
    
    # Contar cuántas filas hay para procesar
    pending_rows = df[(df["ENVAR CORREO"] == "SI") & (df["ESTADO"] == "SIN PROCESAR")]
    logger.info(f"Se encontraron {len(pending_rows)} facturas para enviar correo")
    
    # Procesar cada fila pendiente
    processed_count = 0
    for index, row in pending_rows.iterrows():
        # Información detallada de la fila a procesar
        cliente = row.get("Cliente", "N/A")
        ruc = row.get("RUC", "N/A")
        factura = row.get("N° de Comprobante", "N/A")
        
        logger.info(f"===== Procesando factura {processed_count+1}/{len(pending_rows)} =====")
        logger.info(f"Fila #{index} - Cliente: {cliente}, RUC: {ruc}, Factura: {factura}")
        
        # Mostrar información detallada para diagnóstico
        logger.debug(f"Datos completos de la fila: {dict(row)}")
        
        # Primero actualizar estado a "En Proceso"
        if not update_excel_status(index, "En Proceso"):
            logger.warning(f"No se pudo actualizar el estado a 'En Proceso' para la fila {index}")
            continue
        
        # Ruta de adjuntos formateada
        # Quitar puntos del nombre del cliente y empresa (ej: "INDRA PERU S.A." -> "INDRA PERU SA")
        cliente_sin_puntos = row["Cliente"].replace(".", "")
        empresa_sin_puntos = row["Empresa INDRA/MPS/TCN"].replace(".", "")
        base_path = f'inboxFacturas/{row["Descripción (Primera Fila)"].split("-", 1)[0].strip()}/Organizado/{cliente_sin_puntos}/{empresa_sin_puntos}/{row["Proyecto"]}'
        comprobante_parts = row["N° de Comprobante"].split("-")
        comprobante_num = str(int(comprobante_parts[1]))
        
        # Construir la ruta original
        adjuntos_path = f'{base_path}/{comprobante_parts[0]}-{comprobante_num}'
        logger.info(f"Buscando adjuntos en: {adjuntos_path}")
        
        # Si no existe, intentar con un cero adicional al inicio del número
        if not os.path.isdir(adjuntos_path):
            comprobante_num_con_cero = "0" + comprobante_num
            adjuntos_path_alternativo = f'{base_path}/{comprobante_parts[0]}-{comprobante_num_con_cero}'
            logger.info(f"Directorio original no encontrado. Intentando con: {adjuntos_path_alternativo}")
            
            if os.path.isdir(adjuntos_path_alternativo):
                adjuntos_path = adjuntos_path_alternativo
                logger.info(f"✓ Directorio alternativo encontrado: {adjuntos_path}")
            else:
                logger.warning(f"Tampoco se encontró el directorio alternativo: {adjuntos_path_alternativo}")
        
        # Enviar correo
        email_sent = send_email(row, templates, adjuntos_path)
        
        if email_sent:
            # Actualizar estado a "Enviado" después de enviar el correo exitosamente
            if update_excel_status(index, "Enviado"):
                processed_count += 1
                logger.info(f"✓ COMPLETADO: Correo enviado y estado actualizado para fila {index}")
            else:
                logger.warning(f"No se pudo actualizar el estado a 'Enviado' para la fila {index}")
        else:
            logger.warning(f"❌ FALLIDO: No se pudo enviar el correo para la fila {index}, se mantiene como 'En Proceso'")
        
        logger.info(f"===== Fin del procesamiento de la fila {index} =====\n")
    
    logger.info(f"Proceso completado. {processed_count} de {len(pending_rows)} facturas procesadas completamente")

if __name__ == "__main__":
    logger.info("Iniciando programa de envío de correos")
    process_pending_emails()